from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter
from openpyxl.chart.text import RichText

# =========================
# LOAD FILE
# =========================
wb = load_workbook("temp_schedule.xlsx")

# sheet data chart
ws = wb["CHART_DATA"]

# sheet utama
target_sheet = wb["TIME SCHEDULE"]

# =========================
# CONFIG DARI JS
# =========================
week_start_col = ws["F1"].value

week_end_col = ws["F2"].value

start_data_row = ws["F3"].value

end_data_row = ws["F4"].value

include_progress_chart = bool(ws["F5"].value)

# =========================
# CHART
# =========================
chart = LineChart()

chart.title = None

chart.style = 2

# =========================
# AUTO SIZE CHART
# =========================
total_rows = ((end_data_row + 24) - start_data_row)
total_weeks = (week_end_col - week_start_col) + 1
chart.height = total_rows * 0.60
chart.width = total_weeks * 3.40

# =========================
# TOTAL ROW DATA
# =========================
last_row = ws.max_row

# =========================
# DATA RENCANA
# =========================
data = Reference(
    ws,
    min_col=2,
    max_col=3 if include_progress_chart else 2,
    min_row=2,
    max_row=last_row
)

# =========================
# CATEGORY WEEK
# =========================
cats = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=last_row
)

# =========================
# ADD DATA
# =========================
chart.add_data(
    data,
    titles_from_data=False
)

series1 = chart.series[0]  # rencana

chart.set_categories(cats)

# =========================
# HAPUS LEGEND
# =========================
chart.legend = None

# =========================
# HAPUS AXIS
# =========================
chart.x_axis.delete = True
chart.y_axis.delete = True

# =========================
# SCALE
# =========================
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 100


# =========================
# GARIS MERAH DASHED
# =========================
series1.graphicalProperties = GraphicalProperties()

series1.graphicalProperties.line.solidFill = "CC0000"

series1.graphicalProperties.line.width = 35000

series1.graphicalProperties.line.prstDash = "dash"

# =========================
# MARKER
# =========================
series1.marker.symbol = "circle"

series1.marker.size = 6
# warna titik hitam
series1.graphicalProperties.line.solidFill = "CC0000"

series1.marker.graphicalProperties = GraphicalProperties()

series1.marker.graphicalProperties.solidFill = "000000"

series1.marker.graphicalProperties.line.solidFill = "000000"

series1.smooth = True

# =========================
# DATA LABEL STYLE
# =========================
series1.dLbls = DataLabelList()

series1.dLbls.showVal = True

series1.dLbls.showCatName = False

series1.dLbls.showSerName = False

series1.dLbls.showLegendKey = False

series1.dLbls.numFmt = "0.000"

series1.dLbls.txPr = RichText()

# =========================
# BACKGROUND LABEL
# =========================
series1.dLbls.spPr = GraphicalProperties()

# warna background abu
series1.dLbls.spPr.solidFill = "D9D9D9"

# border hitam
series1.dLbls.spPr.line.solidFill = "000000"

# ketebalan border
# =========================
# GARIS
# =========================
series1.graphicalProperties.line.width = 40000

# =========================
# TITIK
# =========================
series1.marker.size = 6

# =========================
# LABEL BORDER
# =========================
series1.dLbls.spPr.line.width = 20000

# # =========================
# # REALISASI
# # =========================
# series2.graphicalProperties = GraphicalProperties()

# series2.graphicalProperties.line.solidFill = "008000"

# series2.graphicalProperties.line.width = 40000

# series2.marker.symbol = "circle"

# series2.marker.size = 6

# series2.marker.graphicalProperties = GraphicalProperties()

# series2.marker.graphicalProperties.solidFill = "000000"

# series2.marker.graphicalProperties.line.solidFill = "000000"

# series2.smooth = True

# # =========================
# # LABEL
# # =========================
# series2.dLbls = DataLabelList()

# series2.dLbls.showVal = True

# series2.dLbls.showCatName = False

# series2.dLbls.showSerName = False

# series2.dLbls.showLegendKey = False

# series2.dLbls.numFmt = "0.000"

# # posisi label atas
# series2.dLbls.dLblPos = "r"

# # =========================
# # STYLE LABEL
# # =========================
# series2.dLbls.spPr = GraphicalProperties()

# # hijau muda
# series2.dLbls.spPr.solidFill = "C6EFCE"

# # border hijau tua
# series2.dLbls.spPr.line.solidFill = "006100"

# series2.dLbls.spPr.line.width = 20000



chart.graphical_properties = GraphicalProperties(
    noFill=True
)

chart.graphical_properties.line.noFill = True

# =========================
# HAPUS BACKGROUND CHART
# =========================
chart.graphical_properties.noFill = True

# =========================
# HAPUS BACKGROUND PLOT AREA
# =========================
chart.plot_area.graphicalProperties = GraphicalProperties(
    noFill=True
)

# =========================
# HAPUS GRIDLINE
# =========================
chart.y_axis.majorGridlines = None

# =========================
# HAPUS BORDER AXIS
# =========================
chart.x_axis.spPr = GraphicalProperties(
    noFill=True
)

chart.y_axis.spPr = GraphicalProperties(
    noFill=True
)

# =========================
# TARUH CHART KE SCHEDULE
# =========================
# =========================
# POSISI DINAMIS
# =========================
chart_col = get_column_letter(
    week_start_col
)

chart_row = start_data_row + 1

print("TOTAL:", total_rows)

chart.style = 2

chart.roundedCorners = False

target_sheet.add_chart(
    chart,
    f"{chart_col}{chart_row}"
)

# =========================
# SAVE
# =========================
wb.save("final_schedule.xlsx")
